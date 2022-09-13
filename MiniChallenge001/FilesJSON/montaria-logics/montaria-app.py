from enum import Enum
import json
from typing import List
from PyPDF2 import PdfFileReader as pdf_reader
from openpyxl import Workbook, load_workbook as lw


class TipoMoeda(str, Enum):
    cobre = "Cobre"
    prata = "Prata"
    electro = "Electro"
    ouro = "Ouro"
    platina = "Platina"

class Moeda:
    quantidade: int
    tipo: TipoMoeda

    def __init__(self, quantidade: int, tipo: TipoMoeda):
        self.quantidade = quantidade
        self.tipo = tipo

class MontariaAnimal:
    id: int
    nome: str
    preco: Moeda
    delocamento: float
    capacidadeCarga: float

    def __init__(self, id: int, nome: str, preco: Moeda, deslocamento: float, capacidadeCarga: float):
        self.id = id
        self.nome = nome
        self.preco = preco
        self.delocamento = deslocamento
        self.capacidadeCarga = capacidadeCarga

class ItemMontaria:
    id: int
    nome: str
    preco: Moeda
    peso: float

    def __init__(self, id: int, nome: str, preco: Moeda, peso: float):
        self.id = id
        self.nome = nome
        self.preco = preco
        self.peso = peso

class VeiculoAquatico:
    id: int
    nome: str
    preco: Moeda
    velocidade: float

    def __init__(self, id: int, nome: str, preco: Moeda, velocidade: float):
        self.id = id
        self.nome = nome
        self.preco = preco
        self.velocidade = velocidade


def lerPDF():
    reader = pdf_reader("RegrasBasicas-DnD.pdf")
    arquivo = open("pdf_content.txt", "w")
    pagina = reader.pages[155]
    arquivo.write(pagina.extract_text())
    pagina = reader.pages[157]
    arquivo.write(pagina.extract_text())
    arquivo.close()

def criarModeloTabela():
    wb = Workbook()
    wb.create_sheet("Montaria")
    wb.create_sheet("Item Montaria")
    wb.create_sheet("Veículo Aquático")
    wb.save("tabelas.xlsx")
    wb.close()

def criarTabelaMontaria():
    file = open("montaria.txt", "r")
    wb = lw("tabelas.xlsx")
    sheet = wb["Montaria"]

    sheet.cell(1, 1, "ID")
    sheet.cell(1, 2, "NOME")
    sheet.cell(1, 3, "PREÇO : QUANTIDADE")
    sheet.cell(1, 4, "PREÇO : TIPO")
    sheet.cell(1, 5, "DESLOCAMENTO")
    sheet.cell(1, 6, "CAPACIDADE CARGA")
    contador = 1
    while True:
        linha = file.readline()
        if linha == "":
            break
        sheet.cell(contador + 1, 1, contador)
        sheet.cell(contador + 1, 2, linha.strip())
        sheet.cell(contador + 1, 3, int(file.readline().strip()))
        sheet.cell(contador + 1, 4, TipoMoeda(file.readline().strip().replace("po", "Ouro")).value)
        sheet.cell(contador + 1, 5, float(file.readline().replace("m", "").strip()))
        sheet.cell(contador + 1, 6, float(file.readline().strip()))
        contador += 1
        file.readline()
    wb.save("tabelas.xlsx")
    wb.close()

def criarTabelaItemMontaria():
    file = open("item-montaria.txt", "r")
    wb = lw("tabelas.xlsx")
    sheet = wb["Item Montaria"]
    sheet.cell(1, 1, "ID")
    sheet.cell(1, 2, "NOME")
    sheet.cell(1, 3, "PREÇO : QUANTIDADE")
    sheet.cell(1, 4, "PREÇO : TIPO")
    sheet.cell(1, 5, "PESO")
    contador = 1
    while True:
        linha = file.readline()
        if linha == "":
            break
        sheet.cell(contador + 1, 1, contador)
        sheet.cell(contador + 1, 2, linha.strip())
        sheet.cell(contador + 1, 3, int(file.readline().strip()))
        sheet.cell(contador + 1, 4, TipoMoeda(file.readline().strip().replace("po", "Ouro").replace("pc", "Cobre").replace("pp", "Prata")).value)
        sheet.cell(contador + 1, 5, float(file.readline().replace("kg", "").strip().replace(",", ".")))
        file.readline()
        contador += 1
    wb.save("tabelas.xlsx")
    wb.close()

def criarTabelaVeiculo():
    file = open("veiculo-aquatico.txt", "r")
    wb = lw("tabelas.xlsx")
    sheet = wb["Veículo Aquático"]

    sheet.cell(1, 1, "ID")
    sheet.cell(1, 2, "NOME")
    sheet.cell(1, 3, "PRECO : QUANTIDADE")
    sheet.cell(1, 4, "PRECO : TIPO")
    sheet.cell(1, 5, "VELOCIDADE")
    contador = 1
    while True:
        linha = file.readline()
        if linha == "":
            break
        sheet.cell(contador + 1, 1, contador)
        sheet.cell(contador + 1, 2, linha.strip())
        sheet.cell(contador + 1, 3, int(file.readline().strip()))
        sheet.cell(contador + 1, 4, TipoMoeda(file.readline().strip().replace("po", "Ouro")).value)
        sheet.cell(contador + 1, 5, float(file.readline().replace("km/h", "").strip().replace(",", ".")))
        file.readline()
        contador += 1
    wb.save("tabelas.xlsx")
    wb.close()

def criarJsonMontaria():
    file = open("montaria.txt", "r")
    linha = ""
    nome = linha.strip()
    preco_quantidade = int(file.readline().strip())
    preco_tipo = TipoMoeda(file.readline().strip().replace("po", "Ouro"))
    deslocamento = float(file.readline().strip("m").strip())
    capacidade_carga = float(file.readline().strip())

if __name__ == "__main__":
    # lerPDF()
    # criarModeloTabela()
    criarTabelaMontaria()
    criarTabelaItemMontaria()
    criarTabelaVeiculo()
    print("Finished")